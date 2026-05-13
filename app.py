import http.server
import socketserver
import json
import sqlite3
import pandas as pd
import io
import urllib.parse
import time
import psutil
import os
from datetime import datetime

PORT = 5000

# 全局变量存储SQL执行日志
sql_logs = []
start_time = time.time()

def get_db_connection():
    conn = sqlite3.connect('DB/sysdb.db')
    conn.row_factory = sqlite3.Row
    return conn

def log_sql(sql, success=True, error=None):
    """记录SQL执行日志"""
    global sql_logs
    log_entry = {
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'sql': sql,
        'success': success,
        'error': error
    }
    sql_logs.append(log_entry)
    # 只保留最近100条日志
    if len(sql_logs) > 100:
        sql_logs = sql_logs[-100:]

def parse_multipart_form_data(data, boundary):
    """解析multipart/form-data格式的数据"""
    result = {}
    parts = data.split(b'--' + boundary)
    
    for part in parts:
        if not part.strip():
            continue
        
        # 分割头部和内容
        headers_end = part.find(b'\r\n\r\n')
        if headers_end == -1:
            continue
        
        headers = part[:headers_end]
        content = part[headers_end + 4:].rstrip(b'\r\n')
        
        # 解析头部
        headers_str = headers.decode('utf-8')
        disposition_match = None
        for line in headers_str.split('\r\n'):
            if line.startswith('Content-Disposition:'):
                disposition_match = line
                break
        
        if not disposition_match:
            continue
        
        # 提取字段名和文件名
        name = None
        filename = None
        
        # 解析Content-Disposition头部
        parts = disposition_match.split(';')
        for part in parts:
            part = part.strip()
            if part.startswith('name='):
                name = part[5:].strip('"')
            elif part.startswith('filename='):
                filename = part[9:].strip('"')
        
        if not name:
            continue
        
        if filename:
            # 文件字段
            result[name] = {'filename': filename, 'content': content}
        else:
            # 普通字段
            result[name] = content.decode('utf-8')
    
    return result

def get_system_info():
    """获取系统运行信息"""
    try:
        process = psutil.Process(os.getpid())
        
        # 获取内存信息
        memory_info = process.memory_info()
        memory_percent = process.memory_percent()
        
        # 获取CPU信息
        cpu_percent = process.cpu_percent(interval=0.1)
        
        # 获取运行时间
        uptime = time.time() - start_time
        hours, remainder = divmod(int(uptime), 3600)
        minutes, seconds = divmod(remainder, 60)
        uptime_str = f"{hours}小时{minutes}分钟{seconds}秒"
        
        # 获取系统整体信息
        system_memory = psutil.virtual_memory()
        
        # 获取磁盘信息（Windows使用C:\，Linux/Mac使用/）
        try:
            if os.name == 'nt':  # Windows
                disk_usage = psutil.disk_usage('C:\\')
            else:  # Linux/Mac
                disk_usage = psutil.disk_usage('/')
        except:
            disk_usage = None
        
        return {
            'server_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'uptime': uptime_str,
            'process_memory_mb': round(memory_info.rss / 1024 / 1024, 2),
            'process_memory_percent': round(memory_percent, 2),
            'cpu_percent': round(cpu_percent, 2),
            'system_memory_total_gb': round(system_memory.total / 1024 / 1024 / 1024, 2),
            'system_memory_used_gb': round(system_memory.used / 1024 / 1024 / 1024, 2),
            'system_memory_percent': system_memory.percent,
            'disk_total_gb': round(disk_usage.total / 1024 / 1024 / 1024, 2) if disk_usage else 0,
            'disk_used_gb': round(disk_usage.used / 1024 / 1024 / 1024, 2) if disk_usage else 0,
            'disk_percent': disk_usage.percent if disk_usage else 0,
            'sql_logs_count': len(sql_logs)
        }
    except Exception as e:
        return {
            'server_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'uptime': '0小时0分钟0秒',
            'process_memory_mb': 0,
            'process_memory_percent': 0,
            'cpu_percent': 0,
            'system_memory_total_gb': 0,
            'system_memory_used_gb': 0,
            'system_memory_percent': 0,
            'disk_total_gb': 0,
            'disk_used_gb': 0,
            'disk_percent': 0,
            'sql_logs_count': len(sql_logs),
            'error': str(e)
        }

def generate_monitor_html():
    """生成监控页面HTML"""
    return '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SQLite3-SQL-Excel导入分析工具 - 服务器监控</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { 
            font-family: Arial, sans-serif; 
            background: #f0f2f5; 
            padding: 20px; 
        }
        .container { 
            max-width: 1400px; 
            margin: 0 auto; 
        }
        h1 { 
            text-align: center; 
            color: #333; 
            margin-bottom: 30px; 
        }
        .dashboard { 
            display: grid; 
            grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); 
            gap: 20px; 
            margin-bottom: 30px; 
        }
        .card { 
            background: white; 
            border-radius: 8px; 
            padding: 20px; 
            box-shadow: 0 2px 8px rgba(0,0,0,0.1); 
        }
        .card-title { 
            font-size: 16px; 
            font-weight: bold; 
            color: #333; 
            margin-bottom: 15px; 
            padding-bottom: 10px; 
            border-bottom: 2px solid #1890ff; 
        }
        .metric { 
            display: flex; 
            justify-content: space-between; 
            padding: 8px 0; 
            border-bottom: 1px solid #f0f0f0; 
        }
        .metric:last-child { border-bottom: none; }
        .metric-label { color: #666; }
        .metric-value { font-weight: bold; color: #333; }
        .progress-bar { 
            width: 100%; 
            height: 8px; 
            background: #f0f0f0; 
            border-radius: 4px; 
            margin-top: 8px; 
            overflow: hidden;
        }
        .progress-fill { 
            height: 100%; 
            background: #1890ff; 
            border-radius: 4px; 
            transition: width 0.3s ease; 
        }
        .logs-section { 
            background: white; 
            border-radius: 8px; 
            padding: 20px; 
            box-shadow: 0 2px 8px rgba(0,0,0,0.1); 
        }
        .logs-title { 
            font-size: 18px; 
            font-weight: bold; 
            color: #333; 
            margin-bottom: 15px; 
        }
        .log-entry { 
            padding: 12px; 
            margin-bottom: 10px; 
            border-radius: 4px; 
            border-left: 4px solid #1890ff; 
            background: #f6ffed; 
        }
        .log-entry.error { 
            border-left-color: #ff4d4f; 
            background: #fff2f0; 
        }
        .log-time { 
            font-size: 12px; 
            color: #666; 
            margin-bottom: 5px; 
        }
        .log-sql { 
            font-family: monospace; 
            background: #f0f0f0; 
            padding: 8px; 
            border-radius: 4px; 
            word-break: break-all; 
            font-size: 13px;
        }
        .log-error { 
            color: #ff4d4f; 
            margin-top: 5px; 
            font-size: 12px; 
        }
        .refresh-btn { 
            background: #1890ff; 
            color: white; 
            border: none; 
            padding: 8px 20px; 
            border-radius: 4px; 
            cursor: pointer; 
            margin-bottom: 15px; 
        }
        .refresh-btn:hover { background: #40a9ff; }
        .loading { text-align: center; color: #666; padding: 20px; }
        .error-msg { color: #ff4d4f; text-align: center; padding: 20px; }
    </style>
</head>
<body>
    <div class="container">
        <h1>SQLite3-SQL-Excel导入分析工具 - 服务器监控</h1>
        
        <div class="dashboard" id="dashboard">
            <div class="card">
                <div class="card-title">加载中...</div>
                <div class="loading">正在获取系统信息...</div>
            </div>
        </div>
        
        <div class="logs-section">
            <div class="logs-title">SQL执行日志</div>
            <button class="refresh-btn" onclick="refreshData()">刷新数据</button>
            <div id="logs-container">
                <div class="loading">正在加载日志...</div>
            </div>
        </div>
    </div>
    
    <script>
        function refreshData() {
            console.log('开始获取数据...');
            fetch('/api/monitor')
                .then(response => {
                    console.log('收到响应:', response.status);
                    if (!response.ok) {
                        throw new Error('HTTP error! status: ' + response.status);
                    }
                    return response.json();
                })
                .then(data => {
                    console.log('数据获取成功:', data);
                    updateDashboard(data.system_info);
                    updateLogs(data.sql_logs);
                })
                .catch(error => {
                    console.error('获取数据失败:', error);
                    document.getElementById('dashboard').innerHTML = 
                        '<div class="card"><div class="error-msg">获取数据失败: ' + error.message + '</div></div>';
                    document.getElementById('logs-container').innerHTML = 
                        '<div class="error-msg">获取日志失败: ' + error.message + '</div>';
                });
        }
        
        function updateDashboard(info) {
            console.log('更新仪表盘:', info);
            const dashboard = document.getElementById('dashboard');
            
            if (info.error) {
                dashboard.innerHTML = '<div class="card"><div class="error-msg">系统信息获取失败: ' + info.error + '</div></div>';
                return;
            }
            
            dashboard.innerHTML = 
                '<div class="card">' +
                    '<div class="card-title">服务器状态</div>' +
                    '<div class="metric"><span class="metric-label">服务器时间</span><span class="metric-value">' + (info.server_time || '-') + '</span></div>' +
                    '<div class="metric"><span class="metric-label">运行时间</span><span class="metric-value">' + (info.uptime || '-') + '</span></div>' +
                    '<div class="metric"><span class="metric-label">SQL日志数量</span><span class="metric-value">' + (info.sql_logs_count || 0) + '</span></div>' +
                '</div>' +
                '<div class="card">' +
                    '<div class="card-title">进程资源</div>' +
                    '<div class="metric"><span class="metric-label">内存使用</span><span class="metric-value">' + (info.process_memory_mb || 0) + ' MB (' + (info.process_memory_percent || 0) + '%)</span></div>' +
                    '<div class="progress-bar"><div class="progress-fill" style="width: ' + Math.min(info.process_memory_percent || 0, 100) + '%"></div></div>' +
                    '<div class="metric" style="margin-top: 10px;"><span class="metric-label">CPU使用</span><span class="metric-value">' + (info.cpu_percent || 0) + '%</span></div>' +
                    '<div class="progress-bar"><div class="progress-fill" style="width: ' + Math.min(info.cpu_percent || 0, 100) + '%"></div></div>' +
                '</div>' +
                '<div class="card">' +
                    '<div class="card-title">系统内存</div>' +
                    '<div class="metric"><span class="metric-label">总内存</span><span class="metric-value">' + (info.system_memory_total_gb || 0) + ' GB</span></div>' +
                    '<div class="metric"><span class="metric-label">已使用</span><span class="metric-value">' + (info.system_memory_used_gb || 0) + ' GB</span></div>' +
                    '<div class="metric"><span class="metric-label">使用率</span><span class="metric-value">' + (info.system_memory_percent || 0) + '%</span></div>' +
                    '<div class="progress-bar"><div class="progress-fill" style="width: ' + (info.system_memory_percent || 0) + '%"></div></div>' +
                '</div>' +
                '<div class="card">' +
                    '<div class="card-title">磁盘空间</div>' +
                    '<div class="metric"><span class="metric-label">总空间</span><span class="metric-value">' + (info.disk_total_gb || 0) + ' GB</span></div>' +
                    '<div class="metric"><span class="metric-label">已使用</span><span class="metric-value">' + (info.disk_used_gb || 0) + ' GB</span></div>' +
                    '<div class="metric"><span class="metric-label">使用率</span><span class="metric-value">' + (info.disk_percent || 0) + '%</span></div>' +
                    '<div class="progress-bar"><div class="progress-fill" style="width: ' + (info.disk_percent || 0) + '%"></div></div>' +
                '</div>';
        }
        
        function updateLogs(logs) {
            console.log('更新日志:', logs);
            const container = document.getElementById('logs-container');
            if (!logs || logs.length === 0) {
                container.innerHTML = '<p style="text-align: center; color: #666; padding: 20px;">暂无SQL执行记录</p>';
                return;
            }
            
            var html = '';
            for (var i = logs.length - 1; i >= 0; i--) {
                var log = logs[i];
                html += '<div class="log-entry ' + (log.success ? '' : 'error') + '">' +
                    '<div class="log-time">' + (log.timestamp || '-') + '</div>' +
                    '<div class="log-sql">' + escapeHtml(log.sql || '') + '</div>' +
                    (log.error ? '<div class="log-error">错误: ' + escapeHtml(log.error) + '</div>' : '') +
                '</div>';
            }
            container.innerHTML = html;
        }
        
        function escapeHtml(text) {
            if (!text) return '';
            return text.replace(/&/g, '&amp;')
                       .replace(/</g, '&lt;')
                       .replace(/>/g, '&gt;')
                       .replace(/"/g, '&quot;')
                       .replace(/'/g, '&#039;');
        }
        
        // 页面加载完成后自动刷新
        console.log('页面加载完成，开始获取数据...');
        refreshData();
        setInterval(refreshData, 5000);
    </script>
</body>
</html>'''

class MyHTTPRequestHandler(http.server.SimpleHTTPRequestHandler):
    def validate_sql(self, sql):
        """验证SQL语句是否安全，防止SQL注入攻击"""
        if not sql:
            return True
        
        # 转换为小写进行检查
        sql_lower = sql.lower().strip()
        
        # 禁止的危险操作
        dangerous_operations = [
            'drop table',
            'drop database',
            'alter table',
            'create table',
            'truncate table',
            'insert into'
        ]
        
        # 检查是否包含危险操作
        for operation in dangerous_operations:
            if operation in sql_lower:
                return False
        
        # 检查UPDATE操作是否有WHERE子句
        if 'update ' in sql_lower:
            if 'where' not in sql_lower:
                return False
        
        # 检查DELETE操作是否有WHERE子句
        if 'delete from' in sql_lower:
            if 'where' not in sql_lower:
                return False
        
        # 检查是否包含多个语句（分号分隔）
        if ';' in sql:
            return False
        
        # 检查是否包含注释
        if '--' in sql or '/*' in sql:
            return False
        
        return True
    
    def do_GET(self):
        if self.path == '/':
            # 返回监控页面
            self.send_response(200)
            self.send_header('Content-type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(generate_monitor_html().encode('utf-8'))
            return
        elif self.path == '/api/monitor':
            # 返回监控数据
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps({
                'system_info': get_system_info(),
                'sql_logs': sql_logs
            }).encode('utf-8'))
            return
        
        self.send_response(404)
        self.end_headers()
    
    def do_POST(self):
        import json
        if self.path == '/api/execute':
            sql = ''
            page = 1
            page_size = 10
            try:
                content_length = int(self.headers.get('Content-Length', 0))
                if content_length > 0:
                    post_data = self.rfile.read(content_length)
                    data = json.loads(post_data)
                    sql = data.get('sql', '')
                    page = int(data.get('page', 1))
                    page_size = int(data.get('page_size', 10))
                else:
                    sql = ''
                
                # SQL注入防护
                if not self.validate_sql(sql):
                    log_sql(sql, success=False, error='SQL语句包含危险操作')
                    self.send_response(200)
                    self.send_header('Content-type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'success': False,
                        'error': 'SQL语句包含危险操作，请检查后重试'
                    }).encode('utf-8'))
                    return
                
                conn = get_db_connection()
                cursor = conn.cursor()
                
                # 记录执行开始时间
                import time
                start_time = time.time()
                
                cursor.execute(sql)
                
                # 计算执行时间
                execution_time = f"{time.time() - start_time:.4f} 秒"
                
                if sql.strip().lower().startswith('select') or sql.strip().lower().startswith('pragma'):
                    # 执行原始SQL获取所有结果，计算总记录数
                    cursor.execute(sql)
                    total_results = cursor.fetchall()
                    total_count = len(total_results)
                    
                    # 计算分页偏移量
                    offset = (page - 1) * page_size
                    
                    # 构造带分页的SQL语句
                    if sql.strip().lower().startswith('select'):
                        # 对于SELECT语句，添加LIMIT和OFFSET
                        paginated_sql = f"{sql} LIMIT {page_size} OFFSET {offset}"
                        cursor.execute(paginated_sql)
                        results = cursor.fetchall()
                    else:
                        # 对于PRAGMA语句，不进行分页
                        results = total_results
                        paginated_sql = sql
                    
                    if results:
                        columns = [description[0] for description in cursor.description]
                        data = []
                        for row in results:
                            data.append(dict(zip(columns, row)))
                        conn.close()
                        log_sql(sql, success=True)
                        self.send_response(200)
                        self.send_header('Content-type', 'application/json')
                        self.send_header('Access-Control-Allow-Origin', '*')
                        self.end_headers()
                        self.wfile.write(json.dumps({
                            'success': True,
                            'data': data,
                            'columns': columns,
                            'total_count': total_count,
                            'page': page,
                            'page_size': page_size,
                            'execution_time': execution_time
                        }).encode('utf-8'))
                        return
                    else:
                        conn.close()
                        log_sql(sql, success=True)
                        self.send_response(200)
                        self.send_header('Content-type', 'application/json')
                        self.send_header('Access-Control-Allow-Origin', '*')
                        self.end_headers()
                        self.wfile.write(json.dumps({
                            'success': True,
                            'data': [],
                            'columns': [],
                            'total_count': 0,
                            'page': page,
                            'page_size': page_size
                        }).encode('utf-8'))
                        return
                else:
                    # 对于 INSERT、DELETE、UPDATE 等语句，返回影响行数
                    affected_rows = cursor.rowcount
                    conn.commit()
                    conn.close()
                    log_sql(sql, success=True)
                    self.send_response(200)
                    self.send_header('Content-type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'success': True,
                        'affected_rows': affected_rows,
                        'execution_time': execution_time
                    }).encode('utf-8'))
                    return
            except Exception as e:
                log_sql(sql, success=False, error=str(e))
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({
                    'success': False,
                    'error': str(e)
                }).encode('utf-8'))
                return
        
        elif self.path == '/api/export':
            try:
                content_length = int(self.headers.get('Content-Length', 0))
                if content_length > 0:
                    post_data = self.rfile.read(content_length)
                    data = json.loads(post_data)
                    sql = data.get('sql', '')
                else:
                    sql = ''
                
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute(sql)
                
                if cursor.description:
                    columns = [description[0] for description in cursor.description]
                    results = cursor.fetchall()
                    
                    # 导入pandas
                    import pandas as pd
                    
                    # 转换为DataFrame
                    data = []
                    for row in results:
                        data.append(dict(row))
                    df = pd.DataFrame(data)
                    
                    # 导出到Excel
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        # 写入数据，从第一行开始
                        df.to_excel(writer, index=False, sheet_name='结果', startrow=0)
                        
                        # 获取工作表
                        worksheet = writer.sheets['结果']
                        
                        # 导入openpyxl样式
                        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                        from openpyxl.utils import get_column_letter
                        
                        # 表头样式
                        header_font = Font(name='微软雅黑', size=11, bold=True, color='333333')
                        header_fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
                        header_alignment = Alignment(horizontal='center', vertical='center')
                        header_border = Border(
                            top=Side(style='thin', color='DCDFE6'),
                            bottom=Side(style='thin', color='DCDFE6'),
                            left=Side(style='thin', color='DCDFE6'),
                            right=Side(style='thin', color='DCDFE6')
                        )
                        
                        # 数据样式
                        data_font = Font(name='微软雅黑', size=10, color='333333')
                        data_alignment = Alignment(horizontal='left', vertical='center')
                        data_border = Border(
                            top=Side(style='thin', color='F0F0F0'),
                            bottom=Side(style='thin', color='F0F0F0'),
                            left=Side(style='thin', color='F0F0F0'),
                            right=Side(style='thin', color='F0F0F0')
                        )
                        
                        # 设置表头样式
                        for col_num, column_title in enumerate(columns, 1):
                            cell = worksheet.cell(row=1, column=col_num)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                            cell.border = header_border
                        
                        # 设置数据样式
                        for row_num in range(2, len(df) + 2):
                            for col_num in range(1, len(columns) + 1):
                                cell = worksheet.cell(row=row_num, column=col_num)
                                cell.font = data_font
                                cell.alignment = data_alignment
                                cell.border = data_border
                        
                        # 自动调整列宽
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = get_column_letter(column[0].column)
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                        
                        # 调整行高
                        worksheet.row_dimensions[1].height = 25
                        for row in range(2, len(df) + 2):
                            worksheet.row_dimensions[row].height = 20
                    output.seek(0)
                    
                    conn.close()
                    
                    self.send_response(200)
                    self.send_header('Content-type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    self.send_header('Content-Disposition', 'attachment; filename=result.xlsx')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(output.read())
                    return
            except Exception as e:
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({
                    'success': False,
                    'error': str(e)
                }).encode('utf-8'))
                return
        
        elif self.path == '/api/analyze_excel':
            try:
                content_type = self.headers.get('Content-Type', '')
                if not content_type.startswith('multipart/form-data'):
                    self.send_response(200)
                    self.send_header('Content-type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'success': False,
                        'error': 'Content-Type must be multipart/form-data'
                    }).encode('utf-8'))
                    return
                
                # 提取boundary
                boundary = None
                for part in content_type.split(';'):
                    part = part.strip()
                    if part.startswith('boundary='):
                        boundary = part[9:].strip('"').encode('utf-8')
                        break
                
                if not boundary:
                    self.send_response(200)
                    self.send_header('Content-type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'success': False,
                        'error': 'Boundary not found in Content-Type header'
                    }).encode('utf-8'))
                    return
                
                # 读取请求体
                content_length = int(self.headers.get('Content-Length', 0))
                if content_length > 0:
                    post_data = self.rfile.read(content_length)
                else:
                    post_data = b''
                
                # 解析表单数据
                form_data = parse_multipart_form_data(post_data, boundary)
                
                # 获取文件
                file_item = form_data.get('file')
                if not file_item:
                    self.send_response(200)
                    self.send_header('Content-type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'success': False,
                        'error': 'File is required'
                    }).encode('utf-8'))
                    return
                
                # 读取Excel文件，获取列名
                file_content = file_item['content']
                import pandas as pd
                df = pd.read_excel(io.BytesIO(file_content))
                columns = [str(col) for col in df.columns]
                
                # 打印列名，用于调试
                print(f"Excel列名: {columns}")
                
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({
                    'success': True,
                    'columns': columns
                }).encode('utf-8'))
                return
            except Exception as e:
                log_sql('Excel analysis failed', success=False, error=str(e))
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({
                    'success': False,
                    'error': str(e)
                }).encode('utf-8'))
                return
        
        elif self.path == '/api/import':
            try:
                content_type = self.headers.get('Content-Type', '')
                if not content_type.startswith('multipart/form-data'):
                    self.send_response(200)
                    self.send_header('Content-type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'success': False,
                        'error': 'Content-Type must be multipart/form-data'
                    }).encode('utf-8'))
                    return
                
                # 提取boundary
                boundary = None
                for part in content_type.split(';'):
                    part = part.strip()
                    if part.startswith('boundary='):
                        boundary = part[9:].strip('"').encode('utf-8')
                        break
                
                if not boundary:
                    self.send_response(200)
                    self.send_header('Content-type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'success': False,
                        'error': 'Boundary not found in Content-Type header'
                    }).encode('utf-8'))
                    return
                
                # 读取请求体
                content_length = int(self.headers.get('Content-Length', 0))
                if content_length > 0:
                    post_data = self.rfile.read(content_length)
                else:
                    post_data = b''
                
                # 解析表单数据
                form_data = parse_multipart_form_data(post_data, boundary)
                
                # 获取文件和表名
                file_item = form_data.get('file')
                table_name = form_data.get('table_name', '')
                import_type = form_data.get('import_type', 'create')
                field_mappings_json = form_data.get('field_mappings', '[]')
                
                # 清理表名
                import re
                table_name = re.sub(r'[^\w]', '_', table_name)
                if table_name and table_name[0].isdigit():
                    table_name = 'table_' + table_name
                
                if not file_item or not table_name:
                    self.send_response(200)
                    self.send_header('Content-type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'success': False,
                        'error': 'File and table name are required'
                    }).encode('utf-8'))
                    return
                
                # 读取Excel文件
                file_content = file_item['content']
                import pandas as pd
                df = pd.read_excel(io.BytesIO(file_content))
                
                # 获取总条目数
                total_rows = len(df)
                
                # 连接数据库
                conn = get_db_connection()
                cursor = conn.cursor()
                
                # 发送SSE响应头
                self.send_response(200)
                self.send_header('Content-type', 'text/event-stream')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Cache-Control', 'no-cache')
                self.send_header('Connection', 'keep-alive')
                self.end_headers()
                
                def send_progress(progress, processed, total):
                    data = json.dumps({
                        'progress': progress,
                        'processed': processed,
                        'total': total,
                        'status': 'processing'
                    })
                    self.wfile.write(f'data: {data}\n\n'.encode('utf-8'))
                    self.wfile.flush()
                
                def send_completion(message, time_taken):
                    data = json.dumps({
                        'progress': 100,
                        'status': 'completed',
                        'message': message,
                        'time_taken': time_taken,
                        'processed': total_rows
                    })
                    self.wfile.write(f'data: {data}\n\n'.encode('utf-8'))
                    self.wfile.flush()
                
                def send_error(error):
                    data = json.dumps({
                        'progress': 0,
                        'status': 'error',
                        'error': error
                    })
                    self.wfile.write(f'data: {data}\n\n'.encode('utf-8'))
                    self.wfile.flush()
                
                # 发送初始进度信息
                send_progress(0, 0, total_rows)
                
                if import_type == 'create':
                    # 新建表模式
                    # 使用openpyxl获取原始数据类型
                    from openpyxl import load_workbook
                    wb = load_workbook(io.BytesIO(file_content), data_only=True)
                    ws = wb.active
                    
                    # 获取表头和数据 - 清理特殊字符
                    headers = []
                    for cell in ws[1]:
                        col_name = str(cell.value).strip()
                        # 清理特殊字符，只保留字母、数字和下划线
                        import re
                        col_name = re.sub(r'[^\w]', '_', col_name)
                        # 确保不以数字开头，如果是则加前缀
                        if col_name and col_name[0].isdigit():
                            col_name = 'col_' + col_name
                        headers.append(col_name)
                    df.columns = headers
                    
                    # 分析每一列的数据类型（检查前100行非空数据）
                    column_types = {}
                    for col_idx, col_name in enumerate(headers, 1):
                        types_found = {'int': 0, 'float': 0, 'date': 0, 'text': 0}
                        sample_count = 0
                        
                        for row_idx in range(2, min(102, ws.max_row + 1)):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            if cell.value is not None:
                                sample_count += 1
                                cell_type = cell.data_type
                                
                                if cell_type == 'n':  # 数字
                                    if isinstance(cell.value, int):
                                        types_found['int'] += 1
                                    else:
                                        types_found['float'] += 1
                                elif cell_type == 'd':  # 日期
                                    types_found['date'] += 1
                                else:  # 文本或其他
                                    types_found['text'] += 1
                        
                        # 根据样本判断字段类型
                        if sample_count == 0:
                            column_types[col_name] = 'TEXT'
                        elif types_found['int'] > types_found['float'] and types_found['int'] > types_found['text']:
                            column_types[col_name] = 'INTEGER'
                        elif types_found['float'] > types_found['text']:
                            column_types[col_name] = 'REAL'
                        elif types_found['date'] > types_found['text']:
                            column_types[col_name] = 'TEXT'  # 日期也存为TEXT
                        else:
                            column_types[col_name] = 'TEXT'
                    
                    # 检查表是否存在
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
                    if cursor.fetchone():
                        # 表存在，删除后重建
                        cursor.execute(f"DROP TABLE IF EXISTS \"{table_name}\"")
                    
                    # 生成CREATE TABLE语句
                    create_table_sql = f"CREATE TABLE \"{table_name}\" ("
                    columns_def = []
                    
                    # 添加系统字段
                    columns_def.append("sysdbID TEXT PRIMARY KEY")
                    columns_def.append("sysdbDate TEXT")
                    
                    for col in headers:
                        col_type = column_types.get(col, 'TEXT')
                        columns_def.append(f"\"{col}\" {col_type}")
                    
                    create_table_sql += ', '.join(columns_def) + ')'
                    cursor.execute(create_table_sql)
                    
                    # 插入数据 - 高性能优化
                    import time
                    all_columns = ['sysdbID', 'sysdbDate'] + headers
                    
                    # 准备数据
                    total_rows = len(df)
                    batch_size = 1000  # 分批处理
                    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    timestamp = str(int(time.time()))
                    processed_rows = 0
                    start_time = time.time()
                    
                    # 开始事务
                    conn.execute('BEGIN TRANSACTION')
                    
                    try:
                        # 分批处理数据
                        for batch_start in range(0, total_rows, batch_size):
                            batch_end = min(batch_start + batch_size, total_rows)
                            batch_df = df.iloc[batch_start:batch_end]
                            batch_values = []
                            
                            for idx, row in batch_df.iterrows():
                                values = []
                                
                                # 优化ID生成
                                sysdb_id = f"{timestamp}{idx}"
                                values.append(sysdb_id)
                                values.append(current_time)
                                
                                for col in headers:
                                    val = row[col]
                                    # 处理NaN值
                                    if pd.isna(val):
                                        values.append(None)
                                    else:
                                        # 简化日期处理
                                        col_type = column_types.get(col, 'TEXT')
                                        if col_type == 'TEXT':
                                            # 只尝试一种通用日期格式
                                            try:
                                                # 使用pd.to_datetime的自动识别能力
                                                parsed_date = pd.to_datetime(val, errors='ignore')
                                                if isinstance(parsed_date, pd.Timestamp):
                                                    formatted_date = parsed_date.strftime('%Y-%m-%d %H:%M:%S')
                                                    values.append(formatted_date)
                                                else:
                                                    values.append(val)
                                            except:
                                                values.append(val)
                                        else:
                                            values.append(val)
                                batch_values.append(values)
                            
                            # 批量插入
                            if batch_values:
                                placeholders = ['?' for _ in range(len(all_columns))]
                                quoted_columns = ['"sysdbID"', '"sysdbDate"'] + [f'"{col}"' for col in headers]
                                insert_sql = f"INSERT INTO \"{table_name}\" ({', '.join(quoted_columns)}) VALUES ({', '.join(placeholders)})"
                                cursor.executemany(insert_sql, batch_values)
                                
                                # 更新进度
                                processed_rows += len(batch_values)
                                progress = min(100, int((processed_rows / total_rows) * 100))
                                send_progress(progress, processed_rows, total_rows)
                        
                        # 提交事务
                        conn.commit()
                    except Exception as e:
                        # 回滚事务
                        conn.rollback()
                        raise e
                    
                    time_taken = f"{time.time() - start_time:.2f} 秒"
                    message = f'Excel导入成功，表 {table_name} 已创建并导入 {total_rows} 条数据'
                else:
                    # 增量导入模式
                    import json
                    field_mappings = json.loads(field_mappings_json)
                    
                    # 过滤有效的映射
                    valid_mappings = [m for m in field_mappings if m.get('excelColumn') and m.get('dbField')]
                    if not valid_mappings:
                        send_error('No valid field mappings provided')
                        # 关闭数据库连接
                        conn.close()
                        return
                    
                    # 检查表是否存在
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
                    if not cursor.fetchone():
                        send_error(f'Table {table_name} does not exist')
                        # 关闭数据库连接
                        conn.close()
                        return
                    
                    # 插入数据 - 高性能优化
                    import time
                    
                    # 准备字段列表（不含引号，用于数据处理）
                    db_fields = ['sysdbID', 'sysdbDate']
                    for mapping in valid_mappings:
                        db_fields.append(mapping['dbField'])
                    
                    # 准备带引号的字段列表（用于SQL语句）
                    quoted_db_fields = ['"sysdbID"', '"sysdbDate"']
                    for mapping in valid_mappings:
                        quoted_db_fields.append(f'"{mapping["dbField"]}"')
                    
                    # 准备数据
                    total_rows = 0
                    batch_size = 1000  # 分批处理
                    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    timestamp = str(int(time.time()))
                    processed_rows = 0
                    start_time = time.time()
                    
                    # 开始事务
                    conn.execute('BEGIN TRANSACTION')
                    
                    try:
                        # 分批处理数据
                        for batch_start in range(0, len(df), batch_size):
                            batch_end = min(batch_start + batch_size, len(df))
                            batch_df = df.iloc[batch_start:batch_end]
                            batch_values = []
                            
                            for idx, row in batch_df.iterrows():
                                values = []
                                
                                # 优化ID生成
                                sysdb_id = f"{timestamp}{idx}"
                                values.extend([sysdb_id, current_time])
                                
                                for mapping in valid_mappings:
                                    excel_col = mapping['excelColumn']
                                    
                                    if excel_col in df.columns:
                                        val = row[excel_col]
                                        # 处理NaN值
                                        if pd.isna(val):
                                            values.append(None)
                                        else:
                                            # 简化日期处理
                                            try:
                                                # 使用pd.to_datetime的自动识别能力
                                                parsed_date = pd.to_datetime(val, errors='ignore')
                                                if isinstance(parsed_date, pd.Timestamp):
                                                    formatted_date = parsed_date.strftime('%Y-%m-%d %H:%M:%S')
                                                    values.append(formatted_date)
                                                else:
                                                    values.append(val)
                                            except:
                                                values.append(val)
                                    else:
                                        values.append(None)
                                batch_values.append(values)
                                total_rows += 1
                            
                            # 批量插入
                            if batch_values:
                                placeholders = ['?' for _ in range(len(db_fields))]
                                insert_sql = f"INSERT INTO \"{table_name}\" ({', '.join(quoted_db_fields)}) VALUES ({', '.join(placeholders)})"
                                cursor.executemany(insert_sql, batch_values)
                                
                                # 更新进度
                                processed_rows += len(batch_values)
                                progress = min(100, int((processed_rows / len(df)) * 100))
                                send_progress(progress, processed_rows, len(df))
                        
                        # 提交事务
                        conn.commit()
                    except Exception as e:
                        # 回滚事务
                        conn.rollback()
                        raise e
                    
                    time_taken = f"{time.time() - start_time:.2f} 秒"
                    message = f'Excel增量导入成功，表 {table_name} 已追加导入 {total_rows} 条数据'
                
                conn.close()
                
                log_sql(f"Import Excel to table {table_name}", success=True)
                
                # 发送最终进度更新，确保进度条走到100%
                send_progress(100, total_rows, total_rows)
                
                # 发送完成消息
                send_completion(message, time_taken)
                
                return
            except Exception as e:
                log_sql('Excel导入失败', success=False, error=str(e))
                
                # 确保使用SSE格式发送错误消息
                self.send_response(200)
                self.send_header('Content-type', 'text/event-stream')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Cache-Control', 'no-cache')
                self.send_header('Connection', 'keep-alive')
                self.end_headers()
                
                def send_error(error):
                    data = json.dumps({
                        'progress': 0,
                        'status': 'error',
                        'error': error
                    })
                    self.wfile.write(f'data: {data}\n\n'.encode('utf-8'))
                    self.wfile.flush()
                
                send_error(str(e))
                return
        
        elif self.path == '/api/delete_table':
            try:
                content_length = int(self.headers.get('Content-Length', 0))
                if content_length > 0:
                    post_data = self.rfile.read(content_length)
                    data = json.loads(post_data)
                    table_names = data.get('table_names', [])
                else:
                    table_names = []
                
                if not table_names or not isinstance(table_names, list):
                    self.send_response(200)
                    self.send_header('Content-type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'success': False,
                        'error': 'Table names are required'
                    }).encode('utf-8'))
                    return
                
                # 连接数据库
                conn = get_db_connection()
                cursor = conn.cursor()
                
                deleted_tables = []
                for table_name in table_names:
                    # 检查表是否存在
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
                    if cursor.fetchone():
                        # 删除表
                        cursor.execute(f"DROP TABLE IF EXISTS {table_name}")
                        deleted_tables.append(table_name)
                        log_sql(f"DROP TABLE {table_name}", success=True)
                
                conn.commit()
                conn.close()
                
                if len(deleted_tables) > 0:
                    self.send_response(200)
                    self.send_header('Content-type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'success': True,
                        'message': f'成功删除 {len(deleted_tables)} 个表: {', '.join(deleted_tables)}'
                    }).encode('utf-8'))
                else:
                    self.send_response(200)
                    self.send_header('Content-type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'success': False,
                        'error': '没有找到要删除的表'
                    }).encode('utf-8'))
                return
            except Exception as e:
                log_sql(f"Delete tables {table_names} failed", success=False, error=str(e))
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({
                    'success': False,
                    'error': str(e)
                }).encode('utf-8'))
                return
        
        elif self.path == '/api/vacuum':
            try:
                # 导入time模块
                import time
                
                # 连接数据库
                conn = get_db_connection()
                cursor = conn.cursor()
                
                # 获取数据库文件大小（压缩前）
                db_size_before = os.path.getsize('DB/sysdb.db')
                
                # 执行VACUUM操作
                start_time = time.time()
                cursor.execute("VACUUM")
                conn.commit()
                end_time = time.time()
                
                # 获取数据库文件大小（压缩后）
                db_size_after = os.path.getsize('DB/sysdb.db')
                
                conn.close()
                
                time_taken = round(end_time - start_time, 2)
                size_diff = db_size_before - db_size_after
                
                log_sql("VACUUM database", success=True)
                
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({
                    'success': True,
                    'message': f'数据库压缩成功！',
                    'details': {
                        'time_taken': f'{time_taken} 秒',
                        'size_before': f'{round(db_size_before / 1024, 2)} KB',
                        'size_after': f'{round(db_size_after / 1024, 2)} KB',
                        'space_saved': f'{round(size_diff / 1024, 2)} KB'
                    }
                }).encode('utf-8'))
                return
            except Exception as e:
                log_sql("VACUUM database failed", success=False, error=str(e))
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({
                    'success': False,
                    'error': str(e)
                }).encode('utf-8'))
                return
        
        elif self.path == '/api/backup':
            # 数据库备份
            try:
                import shutil
                
                # 生成备份文件名
                backup_dir = 'backups'
                if not os.path.exists(backup_dir):
                    os.makedirs(backup_dir)
                
                backup_filename = f'sysdb_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.db'
                backup_path = os.path.join(backup_dir, backup_filename)
                
                # 复制数据库文件
                shutil.copy2('DB/sysdb.db', backup_path)
                
                # 读取备份文件大小
                backup_size = os.path.getsize(backup_path)
                
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({
                    'success': True,
                    'message': '数据库备份成功',
                    'backup_file': backup_filename,
                    'backup_size': backup_size
                }).encode('utf-8'))
                return
            except Exception as e:
                log_sql('Database backup failed', success=False, error=str(e))
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({
                    'success': False,
                    'error': str(e)
                }).encode('utf-8'))
                return
        
        elif self.path == '/api/restore':
            # 数据库恢复
            try:
                content_type = self.headers.get('Content-Type', '')
                if not content_type.startswith('multipart/form-data'):
                    self.send_response(200)
                    self.send_header('Content-type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'success': False,
                        'error': 'Content-Type must be multipart/form-data'
                    }).encode('utf-8'))
                    return
                
                # 提取boundary
                boundary = None
                for part in content_type.split(';'):
                    part = part.strip()
                    if part.startswith('boundary='):
                        boundary = part[9:].strip('"').encode('utf-8')
                        break
                
                if not boundary:
                    self.send_response(200)
                    self.send_header('Content-type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'success': False,
                        'error': 'Boundary not found in Content-Type header'
                    }).encode('utf-8'))
                    return
                
                # 读取请求体
                content_length = int(self.headers.get('Content-Length', 0))
                if content_length > 0:
                    post_data = self.rfile.read(content_length)
                else:
                    post_data = b''
                
                # 解析表单数据
                form_data = parse_multipart_form_data(post_data, boundary)
                
                # 获取文件
                file_item = form_data.get('file')
                if not file_item:
                    self.send_response(200)
                    self.send_header('Content-type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(json.dumps({
                        'success': False,
                        'error': 'File is required'
                    }).encode('utf-8'))
                    return
                
                # 保存恢复文件
                restore_path = 'DB/sysdb.db'
                with open(restore_path, 'wb') as f:
                    f.write(file_item['content'])
                
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({
                    'success': True,
                    'message': '数据库恢复成功'
                }).encode('utf-8'))
                return
            except Exception as e:
                log_sql('Database restore failed', success=False, error=str(e))
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({
                    'success': False,
                    'error': str(e)
                }).encode('utf-8'))
                return
        
        self.send_response(404)
        self.end_headers()
    
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

import socket

def get_local_ip():
    """获取本地IP地址"""
    try:
        # 创建一个临时的UDP连接来获取本地IP
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return '127.0.0.1'

if __name__ == '__main__':
    local_ip = get_local_ip()
    with socketserver.TCPServer(('0.0.0.0', PORT), MyHTTPRequestHandler) as httpd:
        print(f'Server running at http://{local_ip}:{PORT}')
        print(f'Server running at http://127.0.0.1:{PORT}')
        print(f'Monitor page: http://{local_ip}:{PORT}/')
        httpd.serve_forever()
